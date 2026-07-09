from flask import Flask, request, jsonify, render_template
from flask_cors import CORS
import openpyxl
import math
import io
import os
import socket
import xmlrpc.client
from datetime import date
from urllib.parse import urlparse

app = Flask(__name__)
CORS(app)

socket.setdefaulttimeout(30)

# ── Helpers ──────────────────────────────────────────────────────────────────

def clean_num(v):
    if v is None: return None
    if isinstance(v, (int, float)): return float(v)
    s = str(v).replace('$','').replace('%','').replace(',','.').strip()
    try: return float(s)
    except: return None

def ns_to_z(ns):
    if ns >= 0.99: return 2.326
    if ns >= 0.98: return 2.054
    if ns >= 0.97: return 1.881
    if ns >= 0.95: return 1.645
    return 1.282

def resolve_params(raw):
    """Normaliza los parámetros globales. Acepta niveles de servicio
    (nsA/nsB/nsC, entre 0 y 1 — también admite 95 como 95%) o directamente
    los valores z (za/zb/zc). kpct en % y cloc en $/m²/mes."""
    raw = raw or {}
    def num(v, d):
        v = clean_num(v)
        return v if v is not None else d
    def lvl(v, d):
        v = num(v, d)
        if v > 1: v = v / 100.0   # tolera 95 en vez de 0.95
        return v
    if any(k in raw for k in ('nsA', 'nsB', 'nsC')):
        za = ns_to_z(lvl(raw.get('nsA'), 0.99))
        zb = ns_to_z(lvl(raw.get('nsB'), 0.97))
        zc = ns_to_z(lvl(raw.get('nsC'), 0.95))
    else:
        za = num(raw.get('za'), 2.326)
        zb = num(raw.get('zb'), 1.881)
        zc = num(raw.get('zc'), 1.645)
    return {
        'za': za, 'zb': zb, 'zc': zc,
        'kpct': num(raw.get('kpct'), 4.0),
        'cloc': num(raw.get('cloc'), 9500),
        'alt':  num(raw.get('alt'), 2.0),
    }

def classify_abc(products):
    total_va = sum(p['D'] * p['pv'] for p in products)
    if total_va == 0:
        for p in products: p['clase'] = 'C'
        return
    sorted_p = sorted(products, key=lambda x: x['D'] * x['pv'], reverse=True)
    cum = 0
    for p in sorted_p:
        cum += p['D'] * p['pv'] / total_va * 100
        p['clase'] = 'A' if cum <= 80 else ('B' if cum <= 95 else 'C')
    clase_map = {p['code']: p['clase'] for p in sorted_p}
    for p in products:
        p['clase'] = clase_map.get(p['code'], 'C')

def calc_eoq(p, params):
    kpct  = params['kpct'] / 100
    c_loc = params['cloc']
    alt   = params['alt'] or 2.0
    z_map = {'A': params['za'], 'B': params['zb'], 'C': params['zc']}
    z     = z_map.get(p['clase'], 1.645)
    ms    = p['months']
    n     = len(ms)
    D     = p['D']
    d_avg = D / 12
    variance  = sum((x - d_avg) ** 2 for x in ms) / (n - 1) if n > 1 else 0
    sigma_m   = math.sqrt(variance)
    cv        = sigma_m / d_avg if d_avg > 0 else 0
    sigma_lt  = sigma_m * math.sqrt(p['lt'] / 30)
    SS        = z * sigma_lt
    m2u = p['vol'] / alt
    h   = p['cu'] * 0.25 + m2u * c_loc * 12
    K   = kpct * d_avg * p['cu']
    EOQ   = math.sqrt(2 * D * K / h) if h > 0 else 0
    n_ped = D / EOQ if EOQ > 0 else 0
    ciclo = 365 / n_ped if n_ped > 0 else 0
    PR    = (D / 365) * p['lt'] + SS
    inv_p = EOQ / 2 + SS
    CIT   = (D / EOQ) * K + inv_p * h if EOQ > 0 else 0
    CIT_b = math.sqrt(2 * D * K * h) if EOQ > 0 else 0
    cSS   = SS * h
    rent  = ((p['pv'] - p['cu']) / p['pv']) * 100 if p['pv'] > 0 else 0
    return {
        'z': round(z, 3), 'sigma_m': round(sigma_m, 1), 'cv': round(cv, 3),
        'SS': round(SS, 1), 'h': round(h, 2), 'K': round(K, 2),
        'EOQ': round(EOQ), 'n_ped': round(n_ped, 2), 'ciclo': round(ciclo, 1),
        'PR': round(PR, 1), 'inv_p': round(inv_p, 1),
        'CIT': round(CIT, 2), 'CIT_b': round(CIT_b, 2),
        'cSS': round(cSS, 2), 'rent': round(rent, 2),
        'val_inv': round(inv_p * p['cu'], 2),
    }

def calc_rotation(p):
    """Rotación de inventario, sell-through, backorders y días de inventario."""
    D     = p['D']
    ms    = p['months']
    inv_p = p.get('inv_p', D / 24)  # inventario promedio estimado si no hay EOQ

    # Rotación = ventas anuales / inventario promedio
    rotacion = D / inv_p if inv_p > 0 else 0

    # Días de inventario (DSI)
    dsi = 365 / rotacion if rotacion > 0 else 0

    # Sell-through rate mensual promedio
    # (ventas / (ventas + stock promedio)) * 100
    stock_prom_mensual = inv_p
    ventas_mes = D / 12
    sell_through = (ventas_mes / (ventas_mes + stock_prom_mensual)) * 100 if (ventas_mes + stock_prom_mensual) > 0 else 0

    # Variabilidad mensual para detectar backorders potenciales
    d_avg = D / 12
    variance = sum((x - d_avg) ** 2 for x in ms) / (len(ms) - 1) if len(ms) > 1 else 0
    sigma_m = math.sqrt(variance)

    # Meses con demanda > inventario promedio (riesgo de backorder)
    meses_riesgo = sum(1 for m in ms if m > inv_p)
    pct_riesgo = meses_riesgo / len(ms) * 100

    # Backorder estimado = suma de excesos sobre el inventario promedio
    backorder_est = sum(max(0, m - inv_p) for m in ms)

    return {
        'rotacion': round(rotacion, 2),
        'dsi': round(dsi, 1),
        'sell_through': round(sell_through, 1),
        'meses_riesgo': meses_riesgo,
        'pct_riesgo': round(pct_riesgo, 1),
        'backorder_est': round(backorder_est, 1),
    }

def calc_forecast(months, periods=6, alpha=None):
    """
    Suavizado exponencial simple (Holt-Winters simple).
    Si alpha=None, lo optimiza minimizando MSE.
    Retorna forecast para los próximos `periods` meses y métricas.
    """
    if len(months) < 3:
        avg = sum(months) / len(months)
        return {
            'forecast': [round(avg, 1)] * periods,
            'alpha': 0.3,
            'mae': 0,
            'mape': 0,
            'fitted': months,
            'trend': 0,
        }

    def ses(data, a):
        fitted = [data[0]]
        for i in range(1, len(data)):
            fitted.append(a * data[i-1] + (1 - a) * fitted[i-1])
        return fitted

    def mse(data, a):
        f = ses(data, a)
        return sum((data[i] - f[i])**2 for i in range(1, len(data))) / (len(data) - 1)

    # Optimizar alpha si no se provee
    if alpha is None:
        best_a, best_mse = 0.1, float('inf')
        for a in [i/10 for i in range(1, 10)]:
            m = mse(months, a)
            if m < best_mse:
                best_mse, best_a = m, a
        alpha = best_a

    fitted = ses(months, alpha)
    last = fitted[-1]

    # Detectar tendencia lineal simple
    n = len(months)
    x_avg = (n - 1) / 2
    y_avg = sum(months) / n
    trend = sum((i - x_avg) * (months[i] - y_avg) for i in range(n)) / \
            sum((i - x_avg)**2 for i in range(n)) if n > 1 else 0

    # Forecast: suavizado + tendencia ajustada
    forecast = []
    val = last
    for i in range(periods):
        val = alpha * months[-1] + (1 - alpha) * val + trend * 0.3
        forecast.append(round(max(0, val), 1))

    # MAE y MAPE
    errors = [abs(months[i] - fitted[i]) for i in range(1, len(months))]
    mae = sum(errors) / len(errors) if errors else 0
    pct_errors = [abs(months[i] - fitted[i]) / months[i] * 100
                  for i in range(1, len(months)) if months[i] > 0]
    mape = sum(pct_errors) / len(pct_errors) if pct_errors else 0

    return {
        'forecast': forecast,
        'alpha': round(alpha, 2),
        'mae': round(mae, 1),
        'mape': round(mape, 1),
        'fitted': [round(f, 1) for f in fitted],
        'trend': round(trend, 2),
    }

# ── Integración Odoo (XML-RPC, solo lectura) ─────────────────────────────────

def odoo_base_url(u):
    """Deja solo esquema y dominio: acepta URLs pegadas del navegador con rutas
    como /odoo, /web o /odoo/action-123."""
    u = (u or '').strip()
    if not u:
        return ''
    if not u.startswith('http'):
        u = 'https://' + u
    p = urlparse(u)
    return (p.scheme or 'https') + '://' + p.netloc

def odoo_creds(body):
    """Credenciales del request; las variables de entorno actúan de default."""
    return {
        'url':  odoo_base_url(body.get('url') or os.environ.get('ODOO_URL', '')),
        'db':   (body.get('db')   or os.environ.get('ODOO_DB', '')).strip(),
        'user': (body.get('user') or os.environ.get('ODOO_USER', '')).strip(),
        'key':  (body.get('key')  or os.environ.get('ODOO_API_KEY', '')).strip(),
    }

def odoo_login(c):
    if not all([c['url'], c['db'], c['user'], c['key']]):
        raise ValueError('Faltan datos de conexión: URL, base de datos, usuario y clave API son obligatorios')
    common = xmlrpc.client.ServerProxy(c['url'] + '/xmlrpc/2/common')
    uid = common.authenticate(c['db'], c['user'], c['key'], {})
    if not uid:
        raise ValueError('Autenticación rechazada: revisá la base de datos, el usuario y la clave API')
    models = xmlrpc.client.ServerProxy(c['url'] + '/xmlrpc/2/object')
    return uid, models

def odoo_err(e):
    if 'CERTIFICATE_VERIFY_FAILED' in str(e):
        return 'Error de certificado SSL al conectar — la red desde donde corre el servidor intercepta el tráfico seguro'
    if isinstance(e, xmlrpc.client.ProtocolError):
        return f'El servidor respondió {e.errcode} en {e.url} — verificá que la URL sea la base del Odoo (ej. https://empresa.odoo.com)'
    if isinstance(e, xmlrpc.client.Fault):
        s = e.faultString or ''
        return 'Odoo respondió con error: ' + (s.strip().splitlines()[-1] if s.strip() else 'desconocido')[:300]
    if isinstance(e, (socket.timeout, TimeoutError)):
        return 'Tiempo de espera agotado conectando con Odoo — revisá la URL'
    if isinstance(e, (ConnectionError, OSError)) and not isinstance(e, ValueError):
        return 'No se pudo conectar con la URL indicada — revisá que sea accesible (ej. https://empresa.odoo.com)'
    return str(e)[:300]

def month_start(d, offset):
    m = d.month - 1 + offset
    return date(d.year + m // 12, m % 12 + 1, 1)

# ── Universo de SKU y explosión de BOM ───────────────────────────────────────
#
# El análisis de EOQ corre únicamente sobre artículos SIMPLES (sin mrp.bom
# activa propia). Los compuestos —con BOM tipo Kit (phantom) o Fabricación
# (normal), tratados igual— no se piden a un proveedor: se ensamblan. No
# aparecen como fila en la tabla; su demanda se redistribuye entre sus
# componentes simples, en cualquier nivel de la estructura.
#
# Supuesto de diseño: el consumo derivado se calcula explotando las VENTAS
# del compuesto contra su BOM vigente, no leyendo el consumo físico histórico
# (stock.move / mrp.production). Para EOQ importa la demanda proyectada de
# cada componente, no el momento exacto del consumo físico pasado (que en
# Fabricación puede anteceder a la venta); mezclar fuentes según el tipo de
# BOM introduciría desfasajes de timing difíciles de conciliar.

def hacer_conversor_uom(uoms):
    """Conversor entre UoM con la fórmula de uom.uom._compute_quantity:
    qty / factor_origen * factor_destino. Si falta información o las
    categorías difieren, devuelve la cantidad sin convertir."""
    def convertir(qty, de_id, a_id):
        if not de_id or not a_id or de_id == a_id:
            return qty
        de, a = uoms.get(de_id), uoms.get(a_id)
        if not de or not a or de['categ'] != a['categ']:
            return qty
        return qty / de['factor'] * a['factor']
    return convertir

def explotar_compuestos(simples, compuestos, boms, lineas, uoms, ptav, ventas):
    """Explosión recursiva de BOM sobre las ventas mensuales de los compuestos.

    Devuelve (consumo, forzados, diag, avisos):
      consumo:  {product_id: [12 meses]} = ventas directas del simple +
                consumo derivado de cada compuesto vendido que lo usa
      forzados: ids clasificados como compuestos (por template) pero sin BOM
                aplicable a la variante: entran como fila propia igual
      diag:     desglose compuesto → componente → cantidad derivada anual
      avisos:   anomalías no fatales (BOM circular, componentes omitidos, …)
    """
    simples_por_id = {p['id']: p for p in simples}
    comp_por_id = {p['id']: p for p in compuestos}
    convertir = hacer_conversor_uom(uoms)

    boms_por_tmpl = {}
    for b in sorted(boms, key=lambda b: (b.get('sequence') or 0, b['id'])):
        if b.get('product_tmpl_id'):
            boms_por_tmpl.setdefault(b['product_tmpl_id'][0], []).append(b)
    lineas_por_bom = {}
    for ln in lineas:
        if ln.get('bom_id'):
            lineas_por_bom.setdefault(ln['bom_id'][0], []).append(ln)

    avisos, omitidos, cache = [], set(), {}

    def uom_stock(prod):
        return prod['uom_id'][0] if prod.get('uom_id') else None

    def bom_de(prod):
        """BOM aplicable a la variante (emula mrp.bom._bom_find, que no es
        invocable por RPC): primero la BOM específica de la variante, después
        la genérica del template, en orden de sequence."""
        if not prod.get('product_tmpl_id'):
            return None
        cands = boms_por_tmpl.get(prod['product_tmpl_id'][0], [])
        for b in cands:
            if b.get('product_id') and b['product_id'][0] == prod['id']:
                return b
        for b in cands:
            if not b.get('product_id'):
                return b
        return None

    def explotar_unidad(prod, visitados):
        """{id_simple: cantidad por 1 unidad (en UoM de stock) del compuesto}.
        None si la variante no tiene BOM aplicable (el llamador la trata como
        simple). Lanza ValueError ante una BOM circular; visitados se copia
        por rama para no bloquear ramas hermanas legítimas."""
        if prod['id'] in visitados:
            raise ValueError(f"BOM circular detectada en el producto {prod.get('default_code') or prod['id']}")
        if prod['id'] in cache:
            return cache[prod['id']]
        bom = bom_de(prod)
        if bom is None:
            return None
        cant_bom = convertir(bom.get('product_qty') or 1.0,
                             bom['product_uom_id'][0] if bom.get('product_uom_id') else None,
                             uom_stock(prod)) or 1.0
        mis_ptav = ptav.get(prod['id'], set())
        res = {}
        for ln in lineas_por_bom.get(bom['id'], []):
            if not ln.get('product_id'):
                continue
            ln_ptav = set(ln.get('bom_product_template_attribute_value_ids') or [])
            if ln_ptav and not ln_ptav <= mis_ptav:
                continue  # línea condicionada a otra variante del compuesto
            comp_id = ln['product_id'][0]
            comp = comp_por_id.get(comp_id) or simples_por_id.get(comp_id)
            if comp is None:
                omitidos.add(ln['product_id'][1])  # no almacenable o archivado
                continue
            cant = convertir(ln.get('product_qty') or 0.0,
                             ln['product_uom_id'][0] if ln.get('product_uom_id') else None,
                             uom_stock(comp)) / cant_bom
            if comp_id in comp_por_id:
                sub = explotar_unidad(comp, visitados | {prod['id']})
                if sub is None:
                    res[comp_id] = res.get(comp_id, 0.0) + cant  # sin BOM: simple
                else:
                    for k, v in sub.items():
                        res[k] = res.get(k, 0.0) + v * cant
            else:
                res[comp_id] = res.get(comp_id, 0.0) + cant
        cache[prod['id']] = res
        return res

    consumo = {pid: list(ventas.get(pid) or [0.0] * 12) for pid in simples_por_id}
    forzados, diag = set(), []

    for cp in compuestos:
        ms = ventas.get(cp['id'])
        if not ms or sum(ms) <= 0:
            continue  # sin ventas en el período: no genera consumo derivado
        try:
            por_unidad = explotar_unidad(cp, set())
        except ValueError as e:
            avisos.append(f'{e} — se omitió su explosión (revisar la BOM en Odoo)')
            continue
        if por_unidad is None:
            fila = consumo.setdefault(cp['id'], [0.0] * 12)
            for i in range(12):
                fila[i] += ms[i]
            forzados.add(cp['id'])
            avisos.append(f"{cp.get('default_code') or cp['id']} tiene BOM en el template pero ninguna aplicable a la variante vendida: se trató como simple")
            continue
        det = []
        for comp_id, q in por_unidad.items():
            fila = consumo.setdefault(comp_id, [0.0] * 12)
            for i in range(12):
                fila[i] += ms[i] * q
            if comp_id in comp_por_id:
                forzados.add(comp_id)
            info = simples_por_id.get(comp_id) or comp_por_id.get(comp_id)
            det.append({'code': info.get('default_code') or f'ID-{comp_id}',
                        'name': info['name'], 'qty': round(sum(ms) * q, 2)})
        diag.append({'code': cp.get('default_code') or f"ID-{cp['id']}",
                     'name': cp['name'], 'D': round(sum(ms), 2),
                     'componentes': sorted(det, key=lambda d: -d['qty'])})

    if omitidos:
        lst = sorted(omitidos)
        avisos.append('Componentes de BOM fuera del análisis (no almacenables o archivados): '
                      + ', '.join(lst[:10]) + ('…' if len(lst) > 10 else ''))
    return consumo, forzados, diag, avisos

@app.route('/api/odoo/test', methods=['POST'])
def odoo_test():
    try:
        c = odoo_creds(request.get_json(silent=True) or {})
        uid, models = odoo_login(c)
        n = models.execute_kw(c['db'], uid, c['key'], 'product.product', 'search_count',
                              [[['sale_ok', '=', True]]])
        return jsonify({'ok': True, 'uid': uid, 'productos_vendibles': n})
    except Exception as e:
        return jsonify({'error': odoo_err(e)}), 400

@app.route('/api/odoo/sync', methods=['POST'])
def odoo_sync():
    try:
        body = request.get_json(silent=True) or {}
        c = odoo_creds(body)
        cat = (body.get('categoria') or '').strip()
        uid, models = odoo_login(c)

        def kw(model, method, *args, **kwargs):
            return models.execute_kw(c['db'], uid, c['key'], model, method, list(args), kwargs)

        avisos = []

        # 1) BOMs activas: definen qué producto es compuesto (Kit y Fabricación
        #    cuentan por igual; ver nota de diseño arriba)
        try:
            boms = kw('mrp.bom', 'search_read', [['active', '=', True]],
                      fields=['product_tmpl_id', 'product_id', 'product_qty',
                              'product_uom_id', 'type', 'sequence'])
        except Exception:
            boms = []
            avisos.append('No se pudo leer mrp.bom (¿módulo Fabricación no instalado?); todos los productos se tratan como simples')
        tmpl_compuestos = sorted({b['product_tmpl_id'][0] for b in boms if b.get('product_tmpl_id')})

        # 2) Campo de "almacenable" según la versión de Odoo (is_storable desde v18)
        almacenable = ([['is_storable', '=', True]]
                       if kw('product.product', 'fields_get', ['is_storable'])
                       else [['type', '=', 'product']])

        CAMPOS = ['default_code', 'name', 'volume', 'list_price', 'standard_price',
                  'qty_available', 'product_tmpl_id', 'uom_id']

        # 3) Universo de artículos simples: todo almacenable sin BOM propia,
        #    tenga o no ventas directas. El filtro de categoría restringe solo
        #    las filas del análisis; los compuestos se traen sin filtro porque
        #    su demanda explotada puede caer en componentes de esta categoría.
        dom = list(almacenable)
        if tmpl_compuestos:
            dom.append(['product_tmpl_id', 'not in', tmpl_compuestos])
        if cat:
            dom.append(['categ_id', 'ilike', cat])
        simples = kw('product.product', 'search_read', dom, fields=CAMPOS)

        compuestos = kw('product.product', 'search_read',
                        [['product_tmpl_id', 'in', tmpl_compuestos]],
                        fields=CAMPOS) if tmpl_compuestos else []
        if not simples and not compuestos:
            extra = f' en la categoría "{cat}"' if cat else ''
            return jsonify({'error': 'No se encontraron productos almacenables' + extra}), 400

        # 4) Estructura de las BOM: líneas, UoM y atributos de variante
        lineas, ptav = [], {}
        if boms:
            campos_linea = ['bom_id', 'product_id', 'product_qty', 'product_uom_id',
                            'bom_product_template_attribute_value_ids']
            try:
                lineas = kw('mrp.bom.line', 'search_read',
                            [['bom_id', 'in', [b['id'] for b in boms]]], fields=campos_linea)
            except Exception:
                lineas = kw('mrp.bom.line', 'search_read',
                            [['bom_id', 'in', [b['id'] for b in boms]]], fields=campos_linea[:4])
            try:
                ptav = {p['id']: set(p.get('product_template_attribute_value_ids') or [])
                        for p in kw('product.product', 'read', [p['id'] for p in compuestos],
                                    fields=['product_template_attribute_value_ids'])}
            except Exception:
                pass  # sin datos de variantes: las líneas condicionadas se incluyen igual

        uoms = {}
        uom_ids = {p['uom_id'][0] for p in simples + compuestos if p.get('uom_id')}
        uom_ids |= {b['product_uom_id'][0] for b in boms if b.get('product_uom_id')}
        uom_ids |= {ln['product_uom_id'][0] for ln in lineas if ln.get('product_uom_id')}
        if uom_ids:
            try:
                for u in kw('uom.uom', 'read', list(uom_ids), fields=['factor', 'category_id']):
                    uoms[u['id']] = {'factor': u.get('factor') or 1.0,
                                     'categ': u['category_id'][0] if u.get('category_id') else None}
            except Exception:
                avisos.append('No se pudieron leer las unidades de medida; se asume la misma UoM en BOM y stock')

        # 5) Ventas por mes (últimos 12 cerrados) de TODOS los productos: los
        #    compuestos se consultan para explotarlos aunque no sean fila.
        #    Ojo: no filtrar por product_id acá — read_group repite el domain
        #    entero en el __domain de cada grupo, y con miles de ids la
        #    respuesta XML-RPC crece a decenas de MB y tumba el server; se
        #    filtra client-side con el diccionario `ventas`.
        pids = [p['id'] for p in simples] + [p['id'] for p in compuestos]
        first = month_start(date.today(), -12)
        ventas = {pid: [0.0] * 12 for pid in pids}
        for i in range(12):
            start, end = month_start(first, i), month_start(first, i + 1)
            groups = kw('sale.report', 'read_group',
                        [['date', '>=', str(start)], ['date', '<', str(end)],
                         ['state', 'not in', ['draft', 'sent', 'cancel']]],
                        ['product_uom_qty'], ['product_id'], lazy=False)
            for g in groups:
                pid = g.get('product_id') and g['product_id'][0]
                if pid in ventas:
                    ventas[pid][i] = float(g.get('product_uom_qty') or 0)

        # 6) Consumo total = ventas directas + demanda derivada por explosión
        consumo, forzados, diag, avisos_exp = explotar_compuestos(
            simples, compuestos, boms, lineas, uoms, ptav, ventas)
        avisos += avisos_exp

        info_por_id = {p['id']: p for p in simples}
        info_por_id.update({p['id']: p for p in compuestos if p['id'] in forzados})
        filas = [(pid, ms) for pid, ms in consumo.items()
                 if sum(ms) > 0 and pid in info_por_id]
        if not filas:
            return jsonify({'error': 'Ningún artículo simple registra consumo (directo ni derivado de compuestos) en los últimos 12 meses'}), 400

        # 7) Lead time del proveedor principal, solo para las filas del análisis
        tmpl_ids = list({info_por_id[pid]['product_tmpl_id'][0] for pid, _ in filas
                         if info_por_id[pid].get('product_tmpl_id')})
        delays = {}
        for j in range(0, len(tmpl_ids), 1000):
            for si in kw('product.supplierinfo', 'search_read',
                         [['product_tmpl_id', 'in', tmpl_ids[j:j + 1000]]],
                         fields=['product_tmpl_id', 'delay'], order='sequence'):
                t = si['product_tmpl_id'][0]
                delays.setdefault(t, si.get('delay') or 0)

        products = []
        for pid, ms in filas:
            p = info_por_id[pid]
            tmpl = p['product_tmpl_id'][0] if p.get('product_tmpl_id') else None
            products.append({
                'code':   p.get('default_code') or f"ID-{pid}",
                'name':   p['name'],
                'vol':    p.get('volume') or 0.05,
                'lt':     delays.get(tmpl) or 10,
                'months': [round(m, 2) for m in ms],
                'D':      round(sum(ms), 2),
                'pv':     p.get('list_price') or 0,
                'cu':     p.get('standard_price') or 0,
                'clase':  'A',
                'stock_actual': p.get('qty_available') or 0,
            })

        params = resolve_params(body.get('params'))
        classify_abc(products)
        results = []
        for p in products:
            r = calc_eoq(p, params)
            p.update(r)
            rot = calc_rotation(p)
            fc  = calc_forecast(p['months'])
            results.append({**p, **rot, 'forecast': fc})
        return jsonify({'params': params, 'products': results, 'origen': c['url'],
                        'total_odoo': len(simples),
                        'compuestos_explotados': len(diag),
                        'diagnostico_compuestos': diag,
                        'avisos': avisos})
    except Exception as e:
        return jsonify({'error': odoo_err(e)}), 400

# ── Routes ────────────────────────────────────────────────────────────────────

@app.route('/')
def index():
    return render_template('index.html')

@app.route('/api/parse-excel', methods=['POST'])
def parse_excel():
    if 'file' not in request.files:
        return jsonify({'error': 'No se recibió ningún archivo'}), 400
    f = request.files['file']
    if not f.filename.endswith(('.xlsx', '.xls')):
        return jsonify({'error': 'El archivo debe ser .xlsx o .xls'}), 400
    try:
        wb = openpyxl.load_workbook(io.BytesIO(f.read()), data_only=True)
        ws = wb[wb.sheetnames[0]]
        ns_a  = clean_num(ws.cell(5, 2).value) or 0.99
        ns_b  = clean_num(ws.cell(6, 2).value) or 0.97
        ns_c  = clean_num(ws.cell(7, 2).value) or 0.95
        kpct  = clean_num(ws.cell(8, 2).value) or 0.04
        cloc  = clean_num(ws.cell(9, 2).value) or 9500
        params = {
            'za': ns_to_z(ns_a), 'zb': ns_to_z(ns_b), 'zc': ns_to_z(ns_c),
            'kpct': round(kpct * 100, 2), 'cloc': cloc, 'alt': 2.0,
        }
        products = []
        for row in range(14, ws.max_row + 1):
            code = ws.cell(row, 1).value
            if not code or str(code).strip() == '': break
            months_raw = [clean_num(ws.cell(row, c).value) for c in range(6, 18)]
            valid  = [m for m in months_raw if m is not None and m > 0]
            avg    = sum(valid) / len(valid) if valid else 0
            months = [m if (m is not None and m > 0) else avg for m in months_raw]
            D_total = clean_num(ws.cell(row, 18).value)
            D = D_total or sum(months)
            products.append({
                'code':   str(code).strip(),
                'name':   str(ws.cell(row, 2).value or '').strip(),
                'vol':    clean_num(ws.cell(row, 3).value) or 0.05,
                'lt':     clean_num(ws.cell(row, 5).value) or 10,
                'months': months,
                'D':      D,
                'pv':     clean_num(ws.cell(row, 19).value) or 0,
                'cu':     clean_num(ws.cell(row, 20).value) or 0,
                'clase':  'A',
            })
        if not products:
            return jsonify({'error': 'No se encontraron productos'}), 400
        classify_abc(products)
        results = []
        for p in products:
            r = calc_eoq(p, params)
            p.update(r)
            rot = calc_rotation(p)
            fc  = calc_forecast(p['months'])
            results.append({**p, **rot, 'forecast': fc})
        return jsonify({'params': params, 'products': results})
    except Exception as e:
        return jsonify({'error': f'Error: {str(e)}'}), 500

@app.route('/api/recalc', methods=['POST'])
def recalc():
    try:
        body     = request.get_json()
        products = body.get('products', [])
        params   = resolve_params(body.get('params'))
        if not products: return jsonify({'error': 'Sin productos'}), 400
        classify_abc(products)
        results = []
        for p in products:
            r = calc_eoq(p, params)
            p.update(r)
            rot = calc_rotation(p)
            fc  = calc_forecast(p['months'])
            results.append({**p, **rot, 'forecast': fc})
        return jsonify({'products': results})
    except Exception as e:
        return jsonify({'error': f'Error: {str(e)}'}), 500

if __name__ == '__main__':
    app.run(debug=True)
