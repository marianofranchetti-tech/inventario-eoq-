# Genera static/plantilla_inventario.xlsx — plantilla de datos para el cliente.
# El layout de la hoja "Datos" es el que consume /api/parse-excel (app.py):
#   B5-B9 parámetros · productos desde la fila 14 (A=código ... T=costo unitario)
# La hoja "Instrucciones" documenta cada campo y su equivalente ERP para la
# futura integración por API.
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

AZUL = "1565C0"
AZUL_CLARO = "E3ECF7"
GRIS = "6B7A8D"
BORDE = Side(style="thin", color="E1E5EB")
BOX = Border(left=BORDE, right=BORDE, top=BORDE, bottom=BORDE)

wb = openpyxl.Workbook()

# ── Hoja 1: Datos (la que lee la API) ────────────────────────────────────────
ws = wb.active
ws.title = "Datos"

ws["A1"] = "PLANTILLA DE DATOS — ANÁLISIS DE INVENTARIOS EOQ"
ws["A1"].font = Font(bold=True, size=14, color=AZUL)
ws["A2"] = "Completar las celdas en blanco. No mover filas ni columnas: la posición es fija. Ver hoja «Instrucciones»."
ws["A2"].font = Font(size=10, color=GRIS)

ws["A4"] = "PARÁMETROS GENERALES"
ws["A4"].font = Font(bold=True, size=11)

params = [
    ("Nivel de servicio Clase A (entre 0 y 1)", 0.99),
    ("Nivel de servicio Clase B (entre 0 y 1)", 0.97),
    ("Nivel de servicio Clase C (entre 0 y 1)", 0.95),
    ("Costo de emisión de pedido (fracción del valor de demanda mensual, ej. 0.04 = 4%)", 0.04),
    ("Costo de depósito ($ por m² por mes)", 9500),
]
for i, (label, val) in enumerate(params):
    r = 5 + i
    ws.cell(r, 1, label).font = Font(size=10)
    c = ws.cell(r, 2, val)
    c.fill = PatternFill("solid", fgColor="FFF9E6")
    c.border = BOX
    c.font = Font(bold=True)

ws["A11"] = "PRODUCTOS — una fila por producto, empezando en la fila 14. No dejar filas vacías intermedias."
ws["A11"].font = Font(bold=True, size=11)
ws["A12"] = "Las filas 14 a 16 son EJEMPLOS: reemplazarlas por los productos reales."
ws["A12"].font = Font(size=10, color="C62828", italic=True)

meses = ["Ene", "Feb", "Mar", "Abr", "May", "Jun", "Jul", "Ago", "Sep", "Oct", "Nov", "Dic"]
headers = (
    ["Código", "Descripción", "Volumen unitario (m³)", "Unidad de medida", "Lead time (días)"]
    + [f"Ventas {m} (u)" for m in meses]
    + ["Demanda anual (u) — opcional", "Precio de venta ($/u)", "Costo unitario ($/u)"]
)
for col, h in enumerate(headers, start=1):
    c = ws.cell(13, col, h)
    c.font = Font(bold=True, size=9, color="FFFFFF")
    c.fill = PatternFill("solid", fgColor=AZUL)
    c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    c.border = BOX
ws.row_dimensions[13].height = 40

ejemplos = [
    ["CEM-025", "Cemento Portland x25kg (EJEMPLO)", 0.02, "bolsa", 7,
     3900, 3600, 4100, 4300, 3800, 3500, 3700, 4000, 4200, 4400, 4600, 4900, None, 8500, 6900],
    ["HIE-008", "Hierro nervado 8mm x12m (EJEMPLO)", 0.011, "barra", 14,
     1200, 1100, 1300, 1250, 1150, 1050, 1180, 1220, 1300, 1350, 1400, 1500, None, 11800, 9400],
    ["ARE-BOL", "Arena gruesa bolsón 1m³ (EJEMPLO)", 1.0, "bolsón", 5,
     800, 760, 820, 850, 780, 730, 770, 810, 840, 880, 900, 950, None, 32000, 26500],
]
for i, fila in enumerate(ejemplos):
    r = 14 + i
    for col, val in enumerate(fila, start=1):
        c = ws.cell(r, col, val)
        c.border = BOX
        c.font = Font(size=10, italic=True, color=GRIS)
        if col >= 3:
            c.alignment = Alignment(horizontal="right")

anchos = [12, 34, 12, 12, 10] + [9] * 12 + [14, 14, 14]
for col, w in enumerate(anchos, start=1):
    ws.column_dimensions[get_column_letter(col)].width = w
ws.freeze_panes = "C14"

# ── Hoja 2: Instrucciones + diccionario de campos ────────────────────────────
wi = wb.create_sheet("Instrucciones")
wi["A1"] = "INSTRUCCIONES Y DICCIONARIO DE CAMPOS"
wi["A1"].font = Font(bold=True, size=14, color=AZUL)

notas = [
    "1. Completar la hoja «Datos» sin mover filas ni columnas: el sistema lee posiciones fijas.",
    "2. Los parámetros generales van en las celdas B5 a B9.",
    "3. Los productos se cargan desde la fila 14, una fila por producto, sin filas vacías intermedias.",
    "4. Las filas de ejemplo (14 a 16) deben reemplazarse por los productos reales.",
    "5. Números sin símbolos: usar 6900 (no $6.900). Decimales con punto o coma.",
    "6. Si falta la venta de algún mes, dejar la celda vacía: el sistema la completa con el promedio.",
    "7. La columna «Demanda anual» es opcional: si queda vacía se calcula como la suma de los 12 meses.",
    "8. Guardar como .xlsx y subirlo con el botón «Importar Excel» de la aplicación.",
]
for i, n in enumerate(notas):
    wi.cell(3 + i, 1, n).font = Font(size=10)

wi["A13"] = "DICCIONARIO DE CAMPOS (referencia para la futura integración por API con el ERP)"
wi["A13"].font = Font(bold=True, size=11)

dic_headers = ["Campo", "Ubicación", "Tipo", "Obligatorio", "Descripción", "Equivalente ERP (referencia Odoo)"]
dic = [
    ("Nivel de servicio A/B/C", "B5:B7", "Decimal 0-1", "Sí", "Probabilidad objetivo de no quebrar stock por clase ABC", "Parámetro de planificación (no nativo)"),
    ("Costo de emisión de pedido", "B8", "Decimal", "Sí", "Fracción del valor de demanda mensual que cuesta emitir y recibir un pedido", "purchase.order — costo administrativo"),
    ("Costo de depósito", "B9", "Número", "Sí", "$ por m² por mes de almacenamiento", "stock.warehouse — costo operativo"),
    ("Código", "Col. A (desde fila 14)", "Texto", "Sí", "SKU único del producto; fila vacía corta la lectura", "product.product.default_code"),
    ("Descripción", "Col. B", "Texto", "Sí", "Nombre comercial del producto", "product.product.name"),
    ("Volumen unitario", "Col. C", "Decimal (m³)", "No (usa 0.05)", "Volumen que ocupa una unidad almacenada", "product.product.volume"),
    ("Unidad de medida", "Col. D", "Texto", "No (informativo)", "Unidad de venta; no afecta el cálculo", "product.product.uom_id"),
    ("Lead time", "Col. E", "Entero (días)", "No (usa 10)", "Días entre emitir el pedido y recibir la mercadería", "product.supplierinfo.delay"),
    ("Ventas mensuales Ene-Dic", "Col. F a Q", "Número (unidades)", "Sí (al menos 1 mes)", "Unidades vendidas por mes de los últimos 12 meses", "sale.report — qty agrupada por mes"),
    ("Demanda anual", "Col. R", "Número (unidades)", "No (suma los meses)", "Total anual si se conoce; pisa la suma de los meses", "sale.report — qty anual"),
    ("Precio de venta", "Col. S", "Número ($/u)", "Sí (para ABC y rentabilidad)", "Precio de venta unitario sin IVA", "product.product.list_price"),
    ("Costo unitario", "Col. T", "Número ($/u)", "Sí", "Costo de reposición unitario sin IVA", "product.product.standard_price"),
]
for col, h in enumerate(dic_headers, start=1):
    c = wi.cell(15, col, h)
    c.font = Font(bold=True, size=9, color="FFFFFF")
    c.fill = PatternFill("solid", fgColor=AZUL)
    c.border = BOX
    c.alignment = Alignment(vertical="center", wrap_text=True)
for i, fila in enumerate(dic):
    for col, val in enumerate(fila, start=1):
        c = wi.cell(16 + i, col, val)
        c.font = Font(size=9.5)
        c.border = BOX
        c.alignment = Alignment(vertical="top", wrap_text=True)
    if i % 2 == 0:
        for col in range(1, 7):
            wi.cell(16 + i, col).fill = PatternFill("solid", fgColor="F5F7FA")

for col, w in enumerate([26, 18, 15, 20, 55, 38], start=1):
    wi.column_dimensions[get_column_letter(col)].width = w

wb.save("static/plantilla_inventario.xlsx")
print("plantilla generada: static/plantilla_inventario.xlsx")
